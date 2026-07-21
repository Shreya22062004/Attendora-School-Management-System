from io import BytesIO
from datetime import date
from calendar import monthrange,month_name
import json
from fastapi import APIRouter,Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4,A3,landscape
from reportlab.platypus import SimpleDocTemplate,Table,TableStyle,Paragraph,Spacer,PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from ..database import get_db
from ..auth import require_school_user
from ..models import School,Student,Attendance,SchoolConfig
from .reports import summary_query,studentwise_query,monthly_matrix,yearly_matrix,day_status
from collections import defaultdict
router=APIRouter(prefix='/exports',tags=['Exports'])
CLASS_ORDER={'UKG/KG2/PP1':0, **{str(i):i for i in range(1,13)}}
GENDER_ORDER={'Girl':0,'Boy':1}
CATEGORY_ORDER={'SC':0,'ST':1,'OBC':2,'GENERAL':3,'UNSPECIFIED':4}

def age_on_sep1(dob, year=None):
 if not dob: return ''
 y=year or date.today().year
 cutoff=date(y,9,1)
 return cutoff.year-dob.year-((cutoff.month,cutoff.day)<(dob.month,dob.day))
def student_obj_key(s):
 return (CLASS_ORDER.get(str(s.class_name),99),GENDER_ORDER.get(s.gender,2),(s.name or '').lower())
def category_group_key(item):
 (cls,cat),_ = item
 return (CLASS_ORDER.get(str(cls),99),CATEGORY_ORDER.get(str(cat).upper(),99),str(cat).upper())
def school(db,sid):return db.get(School,sid)
def hdr(ws,s,title):
 ws.append([s.school_name]);ws.append([s.address]);ws.append([f'UDISE CODE: {s.udise_code}']);ws.append([title]);
 for r in range(1,5):ws.cell(r,1).font=Font(bold=True);ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=12)
def xlsx_response(wb,name):
 b=BytesIO();wb.save(b);b.seek(0);return StreamingResponse(b,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename={name}'})
def style_table(t):t.setStyle(TableStyle([('GRID',(0,0),(-1,-1),.35,colors.grey),('BACKGROUND',(0,0),(-1,0),colors.lightgrey),('FONTSIZE',(0,0),(-1,-1),7),('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
def pdf_response(story,name,pagesize=landscape(A4)):
 b=BytesIO();SimpleDocTemplate(b,pagesize=pagesize,rightMargin=18,leftMargin=18,topMargin=20,bottomMargin=20).build(story);b.seek(0);return StreamingResponse(b,media_type='application/pdf',headers={'Content-Disposition':f'attachment; filename={name}'})
def pdf_head(story,s,title):
 st=getSampleStyleSheet();story += [Paragraph(s.school_name,st['Title']),Paragraph(s.address,st['Normal']),Paragraph(f'UDISE CODE: {s.udise_code}',st['Normal']),Paragraph(title,st['Heading2']),Spacer(1,8)]
def summary_sheet(wb,db,sid,start,end,title):
 s=school(db,sid);ws=wb.active;ws.title='Attendance Summary';hdr(ws,s,title);ws.append([]);ws.append(['Class','Boys Present','Girls Present','Total Present','Absent','Marked','Boys Strength','Girls Strength','Total Strength'])
 for r in summary_query(db,sid,start,end):ws.append([r['class_name'],r['boys_present'],r['girls_present'],r['total_present'],r['total_absent'],r['total_marked'],r['boys_total'],r['girls_total'],r['total_students']])
def date_group_gender_totals(db, sid, year, month):
 start=date(year,month,1); end=date(year,month,monthrange(year,month)[1])
 cfg=db.query(SchoolConfig).filter(SchoolConfig.school_id==sid).first()
 classes=json.loads(cfg.classes_json or '[]') if cfg else []
 groups=json.loads(cfg.dashboard_groups_json or '[]') if cfg else []

 # Use the same group fallback as the dashboard.
 if not groups or (len(groups)==1 and groups[0].get('name')=='All Classes'):
  groups=[
   {'name':'UKG/KG2/PP1','classes':[c for c in classes if str(c)=='UKG/KG2/PP1']},
   {'name':'Classes 1 to 5','classes':[c for c in classes if str(c) in ['1','2','3','4','5']]},
   {'name':'Classes 6 to 8','classes':[c for c in classes if str(c) in ['6','7','8']]},
  ]
  groups=[g for g in groups if g['classes']]

 # Normalize configured class values to strings before matching.
 normalized_groups=[]
 for g in groups:
  normalized_groups.append({
   'name':g.get('name') or 'Group',
   'classes':{str(c).strip() for c in (g.get('classes') or [])}
  })

 rows=(db.query(Attendance.attendance_date,Student.class_name,Student.gender,Attendance.status)
       .join(Student,Attendance.student_id==Student.id)
       .filter(Student.school_id==sid,Attendance.attendance_date.between(start,end))
       .all())

 # Show only dates that actually have attendance records.
 attendance_dates=sorted({r.attendance_date for r in rows})
 counts={}
 for r in rows:
  if str(r.status).strip().lower()!='present':
   continue
  student_class=str(r.class_name).strip()
  gender=str(r.gender or '').strip().lower()
  for g in normalized_groups:
   if student_class in g['classes']:
    key=(r.attendance_date,g['name'])
    if key not in counts:
     counts[key]={'girls':0,'boys':0}
    if gender in ('girl','female'):
     counts[key]['girls']+=1
    elif gender in ('boy','male'):
     counts[key]['boys']+=1

 out=[]
 for current in attendance_dates:
  grand_girls=0
  grand_boys=0
  for g in normalized_groups:
   c=counts.get((current,g['name']),{'girls':0,'boys':0})
   out.append({'date':current,'group_name':g['name'],'girls_present':c['girls'],'boys_present':c['boys']})
   grand_girls += c['girls']
   grand_boys += c['boys']
  out.append({'date':current,'group_name':'GRAND TOTAL','girls_present':grand_girls,'boys_present':grand_boys})
 return out

def classwise_category_summary(db, sid):
    from collections import defaultdict

    rows = (
        db.query(
            Student.class_name,
            Student.category,
            Student.gender
        )
        .filter(
            Student.school_id == sid,
            Student.is_active == True
        )
        .all()
    )

    CATEGORY_MAP = {
        "SC": "SC",
        "ST": "ST",
        "OBC": "OBC",
        "GENERAL": "GENERAL"
    }

    result = defaultdict(
        lambda: defaultdict(
            lambda: {"boys": 0, "girls": 0}
        )
    )

    for r in rows:
        cls = str(r.class_name)

        cat = CATEGORY_MAP.get(
            (r.category or "GENERAL").strip().upper(),
            "GENERAL"
        )

        gender = (r.gender or "").strip().lower()

        if gender in ("boy", "male"):
            result[cls][cat]["boys"] += 1
        elif gender in ("girl", "female"):
            result[cls][cat]["girls"] += 1

    return result

def add_classwise_category_sheet(wb, db, sid):

    ws = wb.create_sheet("Classwise Category Summary")

    s = school(db, sid)

    hdr(ws, s, "Classwise Category Summary")

    ws.append([])

    categories = [
        "SC",
        "ST",
        "OBC",
        "GENERAL"
    ]

    row1 = ["Class"]

    for c in categories:
        row1.extend([c, "", ""])

    ws.append(row1)

    row2 = [""]

    for _ in categories:
        row2.extend(["Boys", "Girls", "Total"])

    ws.append(row2)

    summary = classwise_category_summary(db, sid)

    class_order = [
        "UKG/KG2/PP1",
        "1","2","3","4","5",
        "6","7","8",
        "9","10","11","12"
    ]

    grand = defaultdict(int)

    for cls in class_order:

        row = [cls]

        total_boys = 0
        total_girls = 0

        for cat in categories:

            boys = summary[cls][cat]["boys"]
            girls = summary[cls][cat]["girls"]
            total = boys + girls

            row.extend([boys, girls, total])

            total_boys += boys
            total_girls += girls

            grand[(cat, "boys")] += boys
            grand[(cat, "girls")] += girls

        row.extend([
            total_boys,
            total_girls,
            total_boys + total_girls
        ])

        ws.append(row)

    total_row = ["GRAND TOTAL"]

    grand_boys = 0
    grand_girls = 0

    for cat in categories:

        b = grand[(cat, "boys")]
        g = grand[(cat, "girls")]

        total_row.extend([
            b,
            g,
            b + g
        ])

        grand_boys += b
        grand_girls += g

    total_row.extend([
        grand_boys,
        grand_girls,
        grand_boys + grand_girls
    ])

    ws.append(total_row)

    from openpyxl.utils import get_column_letter

    for column in range(1, ws.max_column + 1):
        max_length = 0

        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=column).value
            if value is not None:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[
            get_column_letter(column)
        ].width = max_length + 3
        
def add_date_group_sheet(wb,db,sid,year,month):
 ws=wb.create_sheet('Date Group Gender Totals')
 hdr(ws,school(db,sid),f'Date-wise Group-wise Gender Present Totals - {month_name[month]} {year}')
 ws.append([])
 ws.append(['Date','Group','Girls Present','Boys Present'])
 for r in date_group_gender_totals(db,sid,year,month):
  ws.append([r['date'].isoformat(),r['group_name'],r['girls_present'],r['boys_present']])
 ws.column_dimensions['A'].width=15;ws.column_dimensions['B'].width=28;ws.column_dimensions['C'].width=16;ws.column_dimensions['D'].width=16
 ws.freeze_panes='A7'

def student_sheet(wb,db,sid,start,end):
 ws=wb.create_sheet('Student-wise Attendance');s=school(db,sid);hdr(ws,s,'Student-wise Attendance');ws.append([]);ws.append(['Class','Student Name','Gender','Present','Absent','Working Days','Attendance %'])
 for r in studentwise_query(db,sid,start,end):ws.append([r['class_name'],r['student_name'],r['gender'],r['present'],r['absent'],r['total_marked_days'],r['percentage']])
 set_attendance_excel_widths(ws)
 
def set_attendance_excel_widths(ws, register_type=None, total_days=None):
    # Common columns
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14

    if register_type == 'monthly':
        days = total_days or 31

        # Date columns
        for col_num in range(4, 4 + days):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 6

        # Summary columns after actual last day
        summary_start = 4 + days

        ws.column_dimensions[
            get_column_letter(summary_start)
        ].width = 12

        ws.column_dimensions[
            get_column_letter(summary_start + 1)
        ].width = 12

        ws.column_dimensions[
            get_column_letter(summary_start + 2)
        ].width = 16

        ws.column_dimensions[
            get_column_letter(summary_start + 3)
        ].width = 16

    elif register_type == 'yearly':
        # January to December
        for col_num in range(4, 16):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 14

        # Summary columns
        ws.column_dimensions['P'].width = 12
        ws.column_dimensions['Q'].width = 12
        ws.column_dimensions['R'].width = 16
        ws.column_dimensions['S'].width = 16

    else:
        # Student-wise Attendance
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 16
        ws.column_dimensions['G'].width = 16
        
@router.get('/daily.xlsx')
def daily_x(report_date:date,u=Depends(require_school_user),db:Session=Depends(get_db)):
 wb=Workbook();summary_sheet(wb,db,u.school_id,report_date,report_date,f'Daily Attendance Report - {report_date}');student_sheet(wb,db,u.school_id,report_date,report_date);ws=wb.active;st=day_status(db,u.school_id,report_date);ws.insert_rows(6);ws['A6']=f"Day Status: {st['day_status']} - {st['reason']}";return xlsx_response(wb,f'daily-{report_date}.xlsx')
@router.get('/monthly.xlsx')
def monthly_x(year:int,month:int,u=Depends(require_school_user),db:Session=Depends(get_db)):
 start=date(year,month,1);end=date(year,month,monthrange(year,month)[1]);wb=Workbook();summary_sheet(wb,db,u.school_id,start,end,f'Monthly Attendance Report - {month_name[month]} {year}');student_sheet(wb,db,u.school_id,start,end);ws=wb.create_sheet('Date-wise Register');hdr(ws,school(db,u.school_id),f'Date-wise Register - {month_name[month]} {year}');ws.append([]);last=monthrange(year,month)[1];ws.append(['Class','Student Name','Gender']+[str(i) for i in range(1,last+1)]+['Present','Absent','Working Days','Attendance %'])
 for r in monthly_matrix(db,u.school_id,year,month):ws.append([r['class_name'],r['student_name'],r['gender']]+r['days']+[r['present'],r['absent'],r['working_days'],r['percentage']])
 set_attendance_excel_widths(ws, 'monthly', last)
 ws.freeze_panes = 'D7'
 add_date_group_sheet(wb,db,u.school_id,year,month)
 return xlsx_response(wb, f'monthly-{year}-{month:02d}.xlsx')
@router.get('/yearly.xlsx')
def yearly_x(year:int,u=Depends(require_school_user),db:Session=Depends(get_db)):
 start=date(year,4,1);end=date(year+1,3,31);label=f'{year}-{str(year+1)[-2:]}';months=list(range(4,13))+list(range(1,4));wb=Workbook();summary_sheet(wb,db,u.school_id,start,end,f'Yearly Attendance Report - {label}');student_sheet(wb,db,u.school_id,start,end);ws=wb.create_sheet('Month-wise Register');hdr(ws,school(db,u.school_id),f'Month-wise Register - {label}');ws.append([]);ws.append(['Class','Student Name','Gender']+[month_name[i] for i in months]+['Present','Absent','Working Days','Attendance %'])
 for r in yearly_matrix(db,u.school_id,year):ws.append([r['class_name'],r['student_name'],r['gender']]+r['months']+[r['present'],r['absent'],r['working_days'],r['percentage']])
 set_attendance_excel_widths(ws, 'yearly')
 ws.freeze_panes = 'D7'
 return xlsx_response(wb, f'yearly-{year}.xlsx')

def report_pdf(db,sid,start,end,title,name,matrix=None,matrix_headers=None):
 s=school(db,sid);story=[];pdf_head(story,s,title);rows=[['Class','Boys P','Girls P','Present','Absent','Marked','Boys Str','Girls Str','Strength']]+[[r['class_name'],r['boys_present'],r['girls_present'],r['total_present'],r['total_absent'],r['total_marked'],r['boys_total'],r['girls_total'],r['total_students']] for r in summary_query(db,sid,start,end)];t=Table(rows,repeatRows=1);style_table(t);story += [t,PageBreak()];pdf_head(story,s,'Student-wise Attendance');sw=[['Class','Student Name','Gender','Present','Absent','Marked','%']]+[[r['class_name'],r['student_name'],r['gender'],r['present'],r['absent'],r['total_marked_days'],r['percentage']] for r in studentwise_query(db,sid,start,end)];t=Table(sw,repeatRows=1);style_table(t);story.append(t)
 if matrix is not None:
  story += [PageBreak()];pdf_head(story,s,matrix_headers[0]);rows=[matrix_headers[1]]+matrix;t=Table(rows,repeatRows=1);style_table(t);story.append(t);return pdf_response(story,name,landscape(A3))
 return pdf_response(story,name)
@router.get('/daily.pdf')
def daily_p(report_date:date,u=Depends(require_school_user),db:Session=Depends(get_db)):
 s=school(db,u.school_id);story=[];pdf_head(story,s,f'Daily Attendance Report - {report_date}');st=day_status(db,u.school_id,report_date);story.append(Paragraph(f"Day Status: {st['day_status']} - {st['reason']}",getSampleStyleSheet()['Heading2']));rows=[['Class','Boys P','Girls P','Present','Absent','Marked','Boys Str','Girls Str','Strength']]+[[r['class_name'],r['boys_present'],r['girls_present'],r['total_present'],r['total_absent'],r['total_marked'],r['boys_total'],r['girls_total'],r['total_students']] for r in summary_query(db,u.school_id,report_date,report_date)];t=Table(rows,repeatRows=1);style_table(t);story.append(t);story += [PageBreak()];pdf_head(story,s,'Student-wise Daily Attendance');sw=[['Class','Student Name','Gender','Present','Absent','Marked','%']]+[[r['class_name'],r['student_name'],r['gender'],r['present'],r['absent'],r['total_marked_days'],r['percentage']] for r in studentwise_query(db,u.school_id,report_date,report_date)];tt=Table(sw,repeatRows=1);style_table(tt);story.append(tt);return pdf_response(story,f'daily-{report_date}.pdf')
@router.get('/monthly.pdf')
def monthly_p(year:int,month:int,u=Depends(require_school_user),db:Session=Depends(get_db)):
 start=date(year,month,1);end=date(year,month,monthrange(year,month)[1]);m=monthly_matrix(db,u.school_id,year,month);last=monthrange(year,month)[1]
 headers=['Class','Student Name','Gender']+[str(i) for i in range(1,last+1)]+['P','A','Marked','%']
 rows=[[r['class_name'],r['student_name'],r['gender']]+r['days']+[r['present'],r['absent'],r['working_days'],r['percentage']] for r in m]
 s_obj=school(db,u.school_id);story=[];pdf_head(story,s_obj,f'Monthly Attendance Report - {month_name[month]} {year}')
 summary_rows=[['Class','Boys P','Girls P','Present','Absent','Marked','Boys Str','Girls Str','Strength']]+[[r['class_name'],r['boys_present'],r['girls_present'],r['total_present'],r['total_absent'],r['total_marked'],r['boys_total'],r['girls_total'],r['total_students']] for r in summary_query(db,u.school_id,start,end)]
 t=Table(summary_rows,repeatRows=1);style_table(t);story += [t,PageBreak()];pdf_head(story,s_obj,'Student-wise Attendance')
 sw=[['Class','Student Name','Gender','Present','Absent','Marked','%']]+[[r['class_name'],r['student_name'],r['gender'],r['present'],r['absent'],r['total_marked_days'],r['percentage']] for r in studentwise_query(db,u.school_id,start,end)]
 t=Table(sw,repeatRows=1);style_table(t);story += [t,PageBreak()];pdf_head(story,s_obj,'Date-wise Attendance Register')
 t=Table([headers]+rows,repeatRows=1);style_table(t);story += [t,PageBreak()];pdf_head(story,s_obj,'Date-wise Group-wise Gender Present Totals')
 group_rows=[['Date','Group','Girls Present','Boys Present']]+[[r['date'].isoformat(),r['group_name'],r['girls_present'],r['boys_present']] for r in date_group_gender_totals(db,u.school_id,year,month)]
 t=Table(group_rows,repeatRows=1);style_table(t);story.append(t)
 return pdf_response(story,f'monthly-{year}-{month:02d}.pdf',landscape(A3))

@router.get('/yearly.pdf')
def yearly_p(year:int,u=Depends(require_school_user),db:Session=Depends(get_db)):
 start=date(year,4,1);end=date(year+1,3,31);label=f'{year}-{str(year+1)[-2:]}';months=list(range(4,13))+list(range(1,4));m=yearly_matrix(db,u.school_id,year);headers=['Class','Student Name','Gender']+[month_name[i][:3] for i in months]+['P','A','Marked','%'];rows=[[r['class_name'],r['student_name'],r['gender']]+r['months']+[r['present'],r['absent'],r['working_days'],r['percentage']] for r in m];return report_pdf(db,u.school_id,start,end,f'Yearly Attendance Report - {label}',f'yearly-{year}-{year+1}.pdf',rows,('Month-wise Attendance Register',headers))

@router.get('/students.xlsx')
def students_x(
    u=Depends(require_school_user),
    db: Session = Depends(get_db)
):
    wb = Workbook()

    # -----------------------------
    # SHEET 1: STUDENT DIRECTORY
    # -----------------------------

    ws = wb.active
    ws.title = 'Student Directory'

    hdr(
        ws,
        school(db, u.school_id),
        'Student Directory'
    )

    ws.append([])

    ws.append([
        'Name',
        'Class',
        'Gender',
        'Admission No',
        'PEN Number',
        "Father's Name",
        "Mother's Name",
        'Date of Birth',
        'Age as of 1 Sept',
        'Category',
        'Admission Date'
    ])

    students = sorted(db.query(Student).filter(
        Student.school_id == u.school_id,
        Student.is_active == True
    ).all(), key=student_obj_key)

    for st in students:
        ws.append([
            st.name,
            st.class_name,
            st.gender,
            st.admission_no or '',
            st.pen_number or '',
            st.father_name or '',
            st.mother_name or '',
            str(st.date_of_birth or ''),
            age_on_sep1(st.date_of_birth),
            st.category or 'Unspecified',
            str(st.admission_date or '')
        ])

    # Better widths for Student Directory
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18
    ws.column_dimensions['K'].width = 20

    ws.freeze_panes = 'A7'

    # ----------------------------------------
    # SHEET 2: CLASSWISE CATEGORY STUDENT LIST
    # ----------------------------------------

    ws2 = wb.create_sheet(
        'Classwise Category Lists'
    )

    hdr(
        ws2,
        school(db, u.school_id),
        'Classwise Category Student Lists'
    )

    ws2.append([])

    ws2.append([
        'Class',
        'Category',
        'Student Name',
        'Gender',
        'Admission No',
        'PEN Number',
        "Father's Name",
        "Mother's Name",
        'Date of Birth',
        'Age as of 1 Sept'
    ])

    grouped = {}

    for st in students:
        category = (
            st.category or 'Unspecified'
        ).strip().upper()

        key = (
            st.class_name,
            category
        )

        grouped.setdefault(key, []).append(st)

    for (class_name, category), members in sorted(
        grouped.items(),
        key=category_group_key
    ):
        for st in sorted(members, key=lambda x:(GENDER_ORDER.get(x.gender,2),(x.name or '').lower())):
            ws2.append([
                class_name,
                category,
                st.name,
                st.gender,
                st.admission_no or '',
                st.pen_number or '',
                st.father_name or '',
                st.mother_name or '',
                str(st.date_of_birth or ''),
                age_on_sep1(st.date_of_birth)
            ])

    # Expanded readable column widths
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 28
    ws2.column_dimensions['G'].width = 35
    ws2.column_dimensions['H'].width = 35
    ws2.column_dimensions['I'].width = 18
    ws2.column_dimensions['J'].width = 18

    # Improve row height and alignment
    for row in ws2.iter_rows(min_row=6):
        for cell in row:
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=False
            )

    for row_number in range(6, ws2.max_row + 1):
        ws2.row_dimensions[row_number].height = 22

    ws2.freeze_panes = 'A7'
    ws2.auto_filter.ref = (
        f'A6:J{ws2.max_row}'
    )
    
    add_classwise_category_sheet(
        wb,
        db,
        u.school_id
    )
    
    return xlsx_response(
        wb,
        'student-directory.xlsx'
    )
    
@router.get('/students.pdf')
def students_p(u=Depends(require_school_user),db:Session=Depends(get_db)):
 s=school(db,u.school_id);story=[];pdf_head(story,s,'Student Directory');students=sorted(db.query(Student).filter(Student.school_id==u.school_id,Student.is_active==True).all(),key=student_obj_key);rows=[['Name','Class','Gender','Admission No','PEN Number',"Father's Name","Mother's Name",'Date of Birth','Age as of 1 Sept','Category','Admission Date']]+[[x.name,x.class_name,x.gender,x.admission_no or '',x.pen_number or '',x.father_name or '',x.mother_name or '',str(x.date_of_birth or ''),age_on_sep1(x.date_of_birth),x.category or '',str(x.admission_date or '')] for x in students];t=Table(rows,repeatRows=1);style_table(t);story += [t,PageBreak()];pdf_head(story,s,'Classwise Category Student Lists')
 grouped={}
 for st in students: grouped.setdefault((st.class_name,(st.category or 'Unspecified').upper()),[]).append(st)
 for (cls,cat),members in sorted(grouped.items(),key=category_group_key):
  members=sorted(members,key=lambda x:(GENDER_ORDER.get(x.gender,2),(x.name or '').lower()))
  story.append(Paragraph(f'Class {cls} - {cat}: {len(members)} student(s)',getSampleStyleSheet()['Heading3']))
  rows=[['No.','Student Name','Gender',"Father's Name","Mother's Name",'DOB','Age']]+[[i+1,x.name,x.gender,x.father_name or '',x.mother_name or '',str(x.date_of_birth or ''),age_on_sep1(x.date_of_birth)] for i,x in enumerate(members)];tt=Table(rows,repeatRows=1,colWidths=[30,130,50,130,130,70,40]);style_table(tt);story += [tt,Spacer(1,8)]
 return pdf_response(story,'student-directory.pdf')
